"""
PDF & Excel Report Exporters
"""
import os
from datetime import datetime
from typing import Dict, Any
from fpdf import FPDF, XPos, YPos
from backend.config import settings
from backend.utils.logger import get_logger

logger = get_logger(__name__)


def generate_pdf_report(analysis_data: Dict[str, Any], filename: str) -> str:
    """
    Generate a PDF report from analysis data using FPDF2.
    Returns the absolute path to the generated PDF.
    """
    try:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        # ── Header ────────────────────────────────────────────────────────────
        pdf.set_fill_color(30, 64, 175)  # Royal Blue
        pdf.rect(0, 0, 210, 35, style='F')
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 20)
        pdf.set_y(8)
        pdf.cell(0, 12, "AI Resume & Job Match Analyzer", align="C", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 8, "Analysis Report", align="C", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

        pdf.set_y(40)
        pdf.set_text_color(30, 30, 30)

        # ── Date ──────────────────────────────────────────────────────────────
        pdf.set_font("Helvetica", "I", 9)
        pdf.cell(0, 8, f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", align="R",
                 new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.ln(3)

        # ── Score Summary ─────────────────────────────────────────────────────
        _pdf_section_header(pdf, "Score Summary")

        ats = analysis_data.get("ats_score", {})
        match = analysis_data.get("match_score", {})

        scores = [
            ("Overall ATS Score", ats.get("overall_ats_score", 0)),
            ("Resume Match Score", match.get("match_score", 0)),
            ("Skills Match", ats.get("skills_match_score", 0)),
            ("Experience Match", ats.get("experience_match_score", 0)),
            ("Education Match", ats.get("education_match_score", 0)),
            ("Keyword Match", ats.get("keyword_match_score", 0)),
            ("Formatting Score", ats.get("formatting_score", 0)),
        ]

        for label, score in scores:
            _pdf_score_row(pdf, label, float(score))

        pdf.ln(5)

        # ── Match Category ────────────────────────────────────────────────────
        category = match.get("match_category", "Medium")
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, f"Match Category: {_clean_text(category)}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.ln(3)

        # ── Strengths ─────────────────────────────────────────────────────────
        _pdf_section_header(pdf, "Strengths")
        for strength in match.get("strengths", [])[:6]:
            pdf.set_font("Helvetica", "", 10)
            pdf.cell(5)
            pdf.cell(0, 7, f"- {_clean_text(strength)}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

        # ── Weaknesses ────────────────────────────────────────────────────────
        _pdf_section_header(pdf, "Areas for Improvement")
        for weakness in match.get("weaknesses", [])[:6]:
            pdf.set_font("Helvetica", "", 10)
            pdf.cell(5)
            pdf.cell(0, 7, f"- {_clean_text(weakness)}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

        # ── Missing Skills ────────────────────────────────────────────────────
        missing = ats.get("missing_skills", [])
        if missing:
            _pdf_section_header(pdf, "Missing Skills")
            pdf.set_font("Helvetica", "", 10)
            pdf.multi_cell(0, 7, _clean_text(", ".join(missing[:15])))

        # ── Suggestions ───────────────────────────────────────────────────────
        suggestions = analysis_data.get("suggestions", {})
        rewrite = suggestions.get("resume_rewrite_suggestions", [])
        if rewrite:
            pdf.add_page()
            _pdf_section_header(pdf, "Resume Improvement Suggestions")
            for i, suggestion in enumerate(rewrite[:8], 1):
                pdf.set_font("Helvetica", "", 10)
                pdf.cell(5)
                pdf.multi_cell(0, 7, f"{i}. {_clean_text(suggestion)}")
                pdf.ln(2)

        certs = suggestions.get("recommended_certifications", [])
        if certs:
            _pdf_section_header(pdf, "Recommended Certifications")
            for cert in certs[:6]:
                pdf.set_font("Helvetica", "", 10)
                pdf.cell(5)
                pdf.cell(0, 7, f"- {_clean_text(cert)}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

        # ── Footer ────────────────────────────────────────────────────────────
        pdf.set_y(-20)
        pdf.set_font("Helvetica", "I", 8)
        pdf.set_text_color(120, 120, 120)
        pdf.cell(0, 10, "AI Resume & Job Match Analyzer | MCA Final Year Project | Confidential", align="C")

        # Save
        output_path = os.path.join(settings.reports_path, filename)
        pdf.output(output_path)
        logger.info(f"PDF report generated: {output_path}")
        return output_path

    except Exception as e:
        logger.error(f"PDF generation failed: {e}")
        raise


def _pdf_section_header(pdf, title: str):
    """Helper — render a section header."""
    pdf.ln(3)
    pdf.set_fill_color(241, 245, 249)
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(30, 64, 175)
    pdf.cell(0, 9, f"  {title}", fill=True, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_text_color(30, 30, 30)
    pdf.ln(2)


def _pdf_score_row(pdf, label: str, score: float):
    """Helper — render a labeled score row with progress bar."""
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(70, 7, label)
    # Draw background bar
    pdf.set_fill_color(220, 220, 220)
    pdf.rect(pdf.get_x(), pdf.get_y() + 1, 80, 5, style="F")
    # Draw score bar
    color = _score_color(score)
    pdf.set_fill_color(*color)
    pdf.rect(pdf.get_x(), pdf.get_y() + 1, score * 0.80, 5, style="F")
    pdf.set_x(pdf.get_x() + 82)
    pdf.cell(20, 7, f"{score:.1f}/100", new_x=XPos.LMARGIN, new_y=YPos.NEXT)


def _score_color(score: float):
    if score >= 70:
        return (34, 197, 94)   # Green
    elif score >= 50:
        return (251, 191, 36)  # Yellow
    else:
        return (239, 68, 68)   # Red


def generate_excel_report(analysis_data: Dict[str, Any], filename: str) -> str:
    """
    Generate an Excel report from analysis data using openpyxl.
    Returns the absolute path to the generated Excel file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, Reference

        wb = Workbook()

        # ── Sheet 1: Summary ──────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Analysis Summary"

        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill("solid", fgColor="1E40AF")
        section_font = Font(bold=True, color="1E40AF", size=11)
        section_fill = PatternFill("solid", fgColor="EFF6FF")

        # Title
        ws.merge_cells("A1:D1")
        ws["A1"] = "AI Resume & Job Match Analyzer — Analysis Report"
        ws["A1"].font = Font(bold=True, size=14, color="1E40AF")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
        ws["A2"].font = Font(italic=True, size=9, color="666666")

        # Score table headers
        ws["A4"] = "Metric"
        ws["B4"] = "Score"
        ws["C4"] = "Rating"
        for col in ["A4", "B4", "C4"]:
            ws[col].font = header_font
            ws[col].fill = header_fill
            ws[col].alignment = Alignment(horizontal="center")

        ats = analysis_data.get("ats_score", {})
        match = analysis_data.get("match_score", {})

        rows = [
            ("Overall ATS Score", ats.get("overall_ats_score", 0)),
            ("Resume Match Score", match.get("match_score", 0)),
            ("Skills Match Score", ats.get("skills_match_score", 0)),
            ("Experience Match Score", ats.get("experience_match_score", 0)),
            ("Education Match Score", ats.get("education_match_score", 0)),
            ("Keyword Match Score", ats.get("keyword_match_score", 0)),
            ("Formatting Score", ats.get("formatting_score", 0)),
            ("Soft Skills Score", ats.get("soft_skills_score", 0)),
        ]

        for i, (label, score) in enumerate(rows, start=5):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = round(float(score), 1)
            rating = "Excellent" if score >= 80 else "Good" if score >= 60 else "Average" if score >= 40 else "Needs Work"
            ws[f"C{i}"] = rating

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15

        # ── Sheet 2: Skills ───────────────────────────────────────────────────
        ws2 = wb.create_sheet("Skills Analysis")
        ws2["A1"] = "Matched Skills"
        ws2["B1"] = "Missing Skills"
        for col in ["A1", "B1"]:
            ws2[col].font = header_font
            ws2[col].fill = header_fill

        matched = ats.get("matched_skills", [])
        missing = ats.get("missing_skills", [])
        max_rows = max(len(matched), len(missing))
        for i in range(max_rows):
            if i < len(matched):
                ws2[f"A{i+2}"] = matched[i]
            if i < len(missing):
                ws2[f"B{i+2}"] = missing[i]
        ws2.column_dimensions["A"].width = 25
        ws2.column_dimensions["B"].width = 25

        # ── Sheet 3: Suggestions ──────────────────────────────────────────────
        ws3 = wb.create_sheet("Suggestions")
        suggestions = analysis_data.get("suggestions", {})
        row = 1
        suggestion_groups = [
            ("Resume Rewrite Suggestions", suggestions.get("resume_rewrite_suggestions", [])),
            ("Recommended Certifications", suggestions.get("recommended_certifications", [])),
            ("Suggested Skills", suggestions.get("suggested_skills", [])),
            ("Suggested Projects", suggestions.get("suggested_projects", [])),
            ("Keyword Suggestions", suggestions.get("keyword_suggestions", [])),
        ]
        for group_title, items in suggestion_groups:
            ws3[f"A{row}"] = group_title
            ws3[f"A{row}"].font = section_font
            ws3[f"A{row}"].fill = section_fill
            row += 1
            for item in items:
                ws3[f"A{row}"] = f"• {item}"
                row += 1
            row += 1

        ws3.column_dimensions["A"].width = 80

        output_path = os.path.join(settings.reports_path, filename)
        wb.save(output_path)
        logger.info(f"Excel report generated: {output_path}")
        return output_path

    except Exception as e:
        logger.error(f"Excel generation failed: {e}")
        raise


def _clean_text(txt: str) -> str:
    """Helper to sanitize unicode text to ASCII/Latin-1 characters to prevent FPDF encoding crashes."""
    if not txt:
        return ""
    txt = str(txt)
    replacements = {
        "\u2018": "'", "\u2019": "'",  # curly single quotes
        "\u201c": '"', "\u201d": '"',  # curly double quotes
        "\u2013": "-", "\u2014": "-",  # dashes
        "\u2022": "-",                  # bullet points
        "\u2026": "...",                # ellipsis
    }
    for k, v in replacements.items():
        txt = txt.replace(k, v)
    return txt.encode("latin-1", errors="ignore").decode("latin-1")
